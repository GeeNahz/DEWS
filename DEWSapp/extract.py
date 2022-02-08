import openpyxl as xl
import os
from django.conf import settings




class ExtractData:
    def __init__(self, month, xlsx_file, value=None, yr=None):
        self.month = month
        self.xlsx_file = xlsx_file
        self.value = value
        self.yr = yr

        self.value_select = {
            'mean': 2,
            's': 3,
            'r': 4,
            'b': 5
        }
        self.month_select = {
            "jan": 2,
            "feb": 3,
            "march": 4,
            "april": 5,
            "may": 6,
            "june": 7,
            "july": 8,
            "aug": 9,
            "sept": 10,
            "oct": 11,
            "nov": 12,
            "dec": 13,
        }

        filename = os.path.join(settings.BASE_DIR, 'DEWSapp', 'dewsdocs', f'{self.xlsx_file}.xlsx')

        self.wb = xl.load_workbook(filename) # load a worksheet to the environment
        self.sheet = self.wb['Sheet1'] # specifies the sheet in the worksheet to use


    def value_extract(self):

        no_value = self.value_select[self.value]
        no_month = self.month_select[self.month]

        cell_value = self.sheet.cell(no_value, no_month)

        return cell_value.value

    def yr_extract(self):
        row = 2
        col = 1
        month = self.month_select[self.month]
        while row <= self.sheet.max_row:
            cell = self.sheet.cell(row, col)
            if cell.value == self.yr:
                data = self.sheet.cell(row, month).value
                return data
            row+=1
        return "Nothing found. Please ensure that a valid date was used"

    def save_value(self, data):

        row = 2
        col = 1
        month_col = self.month_select[self.month]

        while row <= self.sheet.max_row:
            cell = self.sheet.cell(row, col)
            if cell.value == self.yr:
                self.sheet.cell(row, month_col).value = data
            row+=1
        self.wb.save(f'DEWSapp\dewsdocs\{self.xlsx_file}.xlsx')
