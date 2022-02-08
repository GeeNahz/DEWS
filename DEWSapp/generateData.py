import openpyxl as xl

from .extract import ExtractData


counter = 0

# def function_itcz(year, cnt=1):
#
#     months = ['jan', 'feb', 'march',
#     'april', 'may', 'june',
#     'july', 'aug', 'sept',
#     'oct', 'nov', 'dec']
#
#     global counter
#
#     while counter < 11:
#         mean_1 = ExtractData(months[cnt], 'PARAitcz10', value='mean').value_extract()
#         mean_0 = ExtractData(months[cnt-1], 'PARAitcz10', value='mean').value_extract()
#         b_0 = ExtractData(months[cnt-1], 'PARAitcz10', value='b').value_extract()
#         rv_1 = ExtractData(months[cnt], 'DATArvITCZ10-1', yr=year).yr_extract()
#         s_1 = ExtractData(months[cnt], 'PARAitcz10', value='s').value_extract()
#         ITCZ_0 = ExtractData(months[cnt-1], 'DATAitcz10', yr=year-1).yr_extract()
#         r_0 = ExtractData(months[cnt-1], 'PARAitcz10', value='r').value_extract()
#
#         ITCZ_1 = mean_1+b_0*(ITCZ_0-mean_0)+rv_1*s_1*((1-(r_0)**2)**0.5)
#
#         ITCZ_1 = round(ITCZ_1, 2)
#
#         ExtractData(months[cnt], 'DATAitcz10', yr=year).save_value(ITCZ_1)
#
#         counter+=1
#
#         function_itcz(year, cnt+1)


def function_sst(year, cnt=1):

    months = ['jan', 'feb', 'march',
    'april', 'may', 'june',
    'july', 'aug', 'sept',
    'oct', 'nov', 'dec']

    global counter

    while counter < 11:
        mean_1 = ExtractData(months[cnt], 'PARAsst', value='mean').value_extract()
        mean_0 = ExtractData(months[cnt-1], 'PARAsst', value='mean').value_extract()
        b_0 = ExtractData(months[cnt-1], 'PARAsst', value='b').value_extract()
        rv_1 = ExtractData(months[cnt], 'DATArvSST-1', yr=year).yr_extract()
        s_1 = ExtractData(months[cnt], 'PARAsst', value='s').value_extract()
        SST_0 = ExtractData(months[cnt-1], 'DATAsst', yr=year-1).yr_extract()
        r_0 = ExtractData(months[cnt-1], 'PARAsst', value='r').value_extract()

        SST_1 = mean_1+b_0*(SST_0-mean_0)+rv_1*s_1*((1-(r_0)**2)**0.5)

        SST_1 = round(SST_1, 2)

        ExtractData(months[cnt], 'DATAsst', yr=year).save_value(SST_1)

        counter+=1

        function_sst(year, cnt+1)



def function_itcz(year, cnt=1):

    months = ['jan', 'feb', 'march',
    'april', 'may', 'june',
    'july', 'aug', 'sept',
    'oct', 'nov', 'dec']

    global counter

    print("ITCZ function started...")

    while counter < 11:
        mean_1 = ExtractData(months[cnt], 'PARAitcz10', value='mean').value_extract()
        mean_0 = ExtractData(months[cnt-1], 'PARAitcz10', value='mean').value_extract()
        b_0 = ExtractData(months[cnt-1], 'PARAitcz10', value='b').value_extract()
        rv_1 = ExtractData(months[cnt], 'DATArvITCZ10-1', yr=year).yr_extract()
        s_1 = ExtractData(months[cnt], 'PARAitcz10', value='s').value_extract()
        SST_0 = ExtractData(months[cnt-1], 'DATAitcz10', yr=year-1).yr_extract()
        r_0 = ExtractData(months[cnt-1], 'PARAitcz10', value='r').value_extract()

        SST_1 = mean_1+b_0*(SST_0-mean_0)+rv_1*s_1*((1-(r_0)**2)**0.5)

        SST_1 = round(SST_1, 2)

        ExtractData(months[cnt], 'DATAitcz10', yr=year).save_value(SST_1)

        counter+=1

        function_itcz(year, cnt+1)
    print("ITCZ function ended...")



















# month = 'jan'
#
# mean_Jan_c = ExtractData(month, 'PARAsst', value='mean')
# mean_Jan = mean_Jan_c.value_extract()
#
# mean_Dec_c = ExtractData(month, 'PARAsst', value='mean')
# mean_Dec = mean_Dec_c.value_extract()
#
# b_Dec_c = ExtractData(month, 'PARAsst', 'b')
# b_Dec = b_Dec_c.value_extract()
#
# ITCZ10D_C = ExtractData(month, 'DATAitcz10', yr=1988-1)
# ITCZ10D = ITCZ10D_C.yr_extract()
#
# rv_Jan = -0.113959645
#
# s_Jan_c = ExtractData(month, 'PARAsst', value='s')
# s_Jan = s_Jan_c.value_extract()
#
# r_Dec_c = ExtractData(month, 'PARAsst', value='r')
# r_Dec = r_Dec_c.value_extract()
#
# ITCZ10J = mean_Jan+b_Dec*(ITCZ10D-mean_Dec)+rv_Jan*s_Jan*((1-(r_Dec)**2)**0.5)
#
# print(ITCZ10J)
#
# def auto_populate(xlsx_file, month, year):
#
#     wb = xl.load_workbook(f'DEWSapp\dewsdocs\{xlsx_file}.xlsx') # load a worksheet to the environment
#     sheet = wb['Sheet1'] # specifies the sheet in the worksheet to use
#
#     month_select = {
#         "jan": 2,
#         "feb": 3,
#         "march": 4,
#         "april": 5,
#         "may": 6,
#         "june": 7,
#         "july": 8,
#         "aug": 9,
#         "sept": 10,
#         "oct": 11,
#         "nov": 12,
#         "dec": 13,
#     }
#
#
#     # loop through the sheet and fill for a year only
#     # if monthval is not more than 13, run this method
#     monthval = month_select[month]
#
#     row = 2
#     counter = 1
#
#     print(ITCZ10J)
#     # print(self.sheet.cell(row, monthval).value)
#     # print(self.yr)
#     while (year is not sheet.cell(row, monthval).value) and (sheet.cell(row, monthval).value is not None):
#         row+=1
#         counter+=1
#
#     # trying if the column and row stuffs work and oh yes it does :)
#     # print(self.sheet.cell(counter+1, monthval).value)
#     # print(counter)
#
#     # now to fill up the empty cells with the right calculated values
#     while monthval <= 13:
#         # check if the value in the cell is empty and if so then fill it with the required value
#         if sheet.cell(counter+1, monthval).value == None:
#             # data_to_fill = self.ExtractData(self.month, 'DATAitcz10', yr=self.yr)
#             print(f"Filling values for month at position {monthval}.")
#         # and if it isn't empty just skip the cell
#         else:
#             print(f"Month at position {monthval} has a value.")
#
#         monthval += 1
    # if it is, end and return a value if required
    #
